"""
update_postal_codes.py
Reads CUSTOMER ADDRESS -SAP.xlsx, derives State from PIN-code prefix,
adds State column to a new Excel file, then injects postal_code into
every customer/location in customer_master.js.

Run with:  uv run --with openpyxl python update_postal_codes.py
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 1. Indian PIN-code → State lookup ─────────────────────────────────────────

# 3-digit prefix overrides (must be checked BEFORE 2-digit)
PIN3 = {
    '160': 'Chandigarh', '161': 'Chandigarh',
    '247': 'Uttarakhand', '248': 'Uttarakhand', '249': 'Uttarakhand',
    '263': 'Uttarakhand', '264': 'Uttarakhand', '265': 'Uttarakhand',
    '500': 'Telangana', '501': 'Telangana', '502': 'Telangana',
    '503': 'Telangana', '504': 'Telangana', '505': 'Telangana',
    '506': 'Telangana', '507': 'Telangana', '508': 'Telangana', '509': 'Telangana',
    '790': 'Arunachal Pradesh', '791': 'Arunachal Pradesh', '792': 'Arunachal Pradesh',
    '793': 'Meghalaya',  '794': 'Meghalaya',
    '795': 'Manipur',   '796': 'Mizoram',
    '797': 'Nagaland',  '798': 'Nagaland',  '799': 'Tripura',
    # Jharkhand overrides inside 81/82 range
    '813': 'Jharkhand', '814': 'Jharkhand', '815': 'Jharkhand',
    '816': 'Jharkhand', '817': 'Jharkhand', '818': 'Jharkhand', '819': 'Jharkhand',
    '820': 'Jharkhand', '821': 'Jharkhand', '822': 'Jharkhand',
    '825': 'Jharkhand', '826': 'Jharkhand', '827': 'Jharkhand',
    '828': 'Jharkhand', '829': 'Jharkhand',
    '830': 'Jharkhand', '831': 'Jharkhand', '832': 'Jharkhand',
    '833': 'Jharkhand', '834': 'Jharkhand', '835': 'Jharkhand',
    '836': 'Jharkhand', '837': 'Jharkhand', '838': 'Bihar',
    '855': 'Jharkhand',
}

# 2-digit prefix fallback
PIN2 = {
    '11': 'Delhi',
    '12': 'Haryana',      '13': 'Haryana',
    '14': 'Punjab',       '15': 'Punjab',       '16': 'Punjab',
    '17': 'Himachal Pradesh',
    '18': 'Jammu & Kashmir',  '19': 'Jammu & Kashmir',
    '20': 'Uttar Pradesh', '21': 'Uttar Pradesh', '22': 'Uttar Pradesh',
    '23': 'Uttar Pradesh', '24': 'Uttar Pradesh', '25': 'Uttar Pradesh',
    '26': 'Uttar Pradesh', '27': 'Uttar Pradesh', '28': 'Uttar Pradesh',
    '30': 'Rajasthan', '31': 'Rajasthan', '32': 'Rajasthan',
    '33': 'Rajasthan', '34': 'Rajasthan',
    '36': 'Gujarat', '37': 'Gujarat', '38': 'Gujarat', '39': 'Gujarat',
    '40': 'Maharashtra', '41': 'Maharashtra', '42': 'Maharashtra',
    '43': 'Maharashtra', '44': 'Maharashtra',
    '45': 'Madhya Pradesh', '46': 'Madhya Pradesh',
    '47': 'Madhya Pradesh', '48': 'Madhya Pradesh',
    '49': 'Chhattisgarh',
    '50': 'Telangana',
    '51': 'Andhra Pradesh', '52': 'Andhra Pradesh', '53': 'Andhra Pradesh',
    '54': 'Karnataka', '55': 'Karnataka', '56': 'Karnataka',
    '57': 'Karnataka', '58': 'Karnataka', '59': 'Karnataka',
    '60': 'Tamil Nadu', '61': 'Tamil Nadu', '62': 'Tamil Nadu',
    '63': 'Tamil Nadu', '64': 'Tamil Nadu', '65': 'Tamil Nadu', '66': 'Tamil Nadu',
    '67': 'Kerala', '68': 'Kerala', '69': 'Kerala',
    '70': 'West Bengal', '71': 'West Bengal', '72': 'West Bengal',
    '73': 'West Bengal', '74': 'West Bengal',
    '75': 'Odisha', '76': 'Odisha', '77': 'Odisha',
    '78': 'Assam', '79': 'Assam',
    '80': 'Bihar', '81': 'Bihar',
    '82': 'Jharkhand', '83': 'Jharkhand',
    '84': 'Bihar', '85': 'Bihar',
    '86': 'Jharkhand', '87': 'Jharkhand',
    '88': 'Bihar', '89': 'Bihar',
}

def state_from_pin(pin):
    p = str(pin).zfill(6) if pin else ''
    if len(p) < 2:
        return ''
    if p[:3] in PIN3:
        return PIN3[p[:3]]
    return PIN2.get(p[:2], '')


# ── 2. Read Excel ──────────────────────────────────────────────────────────────

EXCEL_IN  = r'C:\Users\Admin\Downloads\CUSTOMER ADDRESS -SAP.xlsx'
EXCEL_OUT = r'C:\Users\Admin\Downloads\CUSTOMER ADDRESS -SAP-Updated.xlsx'
MASTER_JS = r'C:\Projects\Costing_App\costing\ravasco-cds\masters\customer_master.js'

wb = openpyxl.load_workbook(EXCEL_IN, data_only=True)
ws = wb.active

excel_rows = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    name = (str(row[1]) if row[1] else '').strip().upper()
    city = (str(row[6]) if row[6] else '').strip().upper()
    raw_pin = row[7]
    pin = str(int(raw_pin)).zfill(6) if raw_pin else ''
    excel_rows.append({'name': name, 'city': city, 'pin': pin})

print(f"Excel rows read: {len(excel_rows)}")

# ── 3. Build (name, city) → pin  AND  city → state from customer_master.js ───

with open(MASTER_JS, 'r', encoding='utf-8') as f:
    master_text = f.read()
    master_lines = master_text.splitlines(keepends=True)

# Collect (name_upper, city_upper) → state  from seed data
name_city_to_state = {}
city_to_state      = {}

for line in master_lines:
    if 'id: "CUS-' not in line:
        continue
    nm = re.search(r'name:\s*"([^"]*)"', line)
    if not nm:
        continue
    cname = nm.group(1).upper()

    if 'locations: []' in line:
        cm = re.search(r',\s*city:\s*"([^"]*)"', line)
        sm = re.search(r',\s*state:\s*"([^"]*)"', line)
        if cm and sm and sm.group(1):
            k = (cname, cm.group(1).upper())
            name_city_to_state[k] = sm.group(1)
            if cm.group(1):
                city_to_state[cm.group(1).upper()] = sm.group(1)
    else:
        for loc in re.finditer(r'\{[^}]*id:\s*"CUS-[^"]*-\d+"[^}]*\}', line):
            ls = loc.group(0)
            cm = re.search(r'city:\s*"([^"]*)"', ls)
            sm = re.search(r'state:\s*"([^"]*)"', ls)
            if cm and sm and sm.group(1):
                k = (cname, cm.group(1).upper())
                name_city_to_state[k] = sm.group(1)
                if cm.group(1):
                    city_to_state[cm.group(1).upper()] = sm.group(1)

print(f"name+city->state entries: {len(name_city_to_state)}")

# ── 4. Build (name, city) → pin  from Excel ───────────────────────────────────
# Only include empty-city rows when that customer has exactly ONE empty-city entry
# (multiple empty-city rows can't be reliably matched to specific locations)
name_city_to_pin = {}
from collections import Counter
empty_city_counts = Counter(r['name'] for r in excel_rows if not r['city'] and r['pin'])

for r in excel_rows:
    if r['pin']:
        name = r['name']
        city = r['city']
        if not city and empty_city_counts[name] > 1:
            continue  # ambiguous — skip, location will get postal_code: null
        name_city_to_pin[(name, city)] = r['pin']

# ── 5. Add State column to Excel & save ──────────────────────────────────────

STATE_COL = ws.max_column + 1

# Header
hdr_cell = ws.cell(row=1, column=STATE_COL, value='State')
hdr_cell.font      = Font(name='Arial', bold=True, color='FFFFFF')
hdr_cell.fill      = PatternFill('solid', start_color='17375E')
hdr_cell.alignment = Alignment(horizontal='center', vertical='center')

matched_state = 0
pin_derived   = 0
empty_state   = 0

for i, r in enumerate(excel_rows):
    row_num  = i + 2
    name     = r['name']
    city     = r['city']
    pin      = r['pin']

    state = (name_city_to_state.get((name, city))
             or city_to_state.get(city)
             or state_from_pin(pin))

    if state:
        if name_city_to_state.get((name, city)):
            matched_state += 1
        else:
            pin_derived += 1
    else:
        empty_state += 1

    ws.cell(row=row_num, column=STATE_COL, value=state or '')

# Make State column width comfortable
from openpyxl.utils import get_column_letter
ws.column_dimensions[get_column_letter(STATE_COL)].width = 22

wb.save(EXCEL_OUT)
print(f"Excel saved → {EXCEL_OUT}")
print(f"  State matched from master : {matched_state}")
print(f"  State from PIN prefix     : {pin_derived}")
print(f"  State empty               : {empty_state}")

# ── 6. Inject postal_code into customer_master.js ────────────────────────────

new_lines = []
pins_added = 0
locs_added = 0

for line in master_lines:
    raw = line  # preserve original if no match
    if 'id: "CUS-' not in line:
        new_lines.append(line)
        continue
    nm = re.search(r'name:\s*"([^"]*)"', line)
    if not nm:
        new_lines.append(line)
        continue

    cname = nm.group(1).upper()

    if 'locations: []' in line:
        # ── Single-site: add postal_code between address and locations ─────────
        cm = re.search(r',\s*city:\s*"([^"]*)"', line)
        city = cm.group(1).upper() if cm else ''
        pin = name_city_to_pin.get((cname, city), '')
        pc  = f'"{pin}"' if pin else 'null'

        def _root_sub(m, _pc=pc):
            return f'{m.group(1)}, postal_code: {_pc},{m.group(2)}'

        new_line = re.sub(
            r'(address:\s*(?:"[^"]*"|null)),(\s*locations:)',
            _root_sub, line
        )
        if new_line != line:
            pins_added += 1
        new_lines.append(new_line)

    else:
        # ── Multi-site ────────────────────────────────────────────────────────
        # a) Add postal_code: null on root (address is always null for multi-site)
        line = re.sub(
            r'(address:\s*null),(\s*locations:)',
            r'\1, postal_code: null,\2',
            line, count=1
        )

        # b) Process each location sub-object
        def _loc_sub(m, _cname=cname):
            loc = m.group(0)
            cm2 = re.search(r'city:\s*"([^"]*)"', loc)
            lc  = cm2.group(1).upper() if cm2 else ''
            p   = name_city_to_pin.get((_cname, lc), '')
            pc2 = f'"{p}"' if p else 'null'

            replaced = re.sub(
                r'(address:\s*(?:"[^"]*"|null))(\s*\})',
                lambda mm, _p=pc2: f'{mm.group(1)}, postal_code: {_p}{mm.group(2)}',
                loc
            )
            return replaced

        new_line = re.sub(r'\{[^}]*id:\s*"CUS-[^"]*-\d+"[^}]*\}', _loc_sub, line)
        locs_added += new_line.count('postal_code:') - line.count('postal_code:')
        new_lines.append(new_line)
        continue

new_content = ''.join(new_lines)

# ── 7. Bump data version to force localStorage refresh ────────────────────────
new_content = new_content.replace(
    "const CUSTOMER_DATA_VERSION = 'v2-cleaned-741';",
    "const CUSTOMER_DATA_VERSION = 'v3-pins-741';"
)

# ── 8. Update CUSTOMER_FIELDS to include postal_code ─────────────────────────
new_content = new_content.replace(
    "  'address',            // Full address (null on parent when locations[] is used)",
    "  'address',            // Full address (null on parent when locations[] is used)\n  'postal_code',        // 6-digit Indian PIN code"
)

# ── 9. Update location objects comment in CUSTOMER_FIELDS ─────────────────────
new_content = new_content.replace(
    "  'locations',          // Array of { id, gst, city, state, address } — empty for single-site",
    "  'locations',          // Array of { id, gst, city, state, address, postal_code } — empty for single-site"
)

# ── 10. Save customer_master.js ────────────────────────────────────────────────
with open(MASTER_JS, 'w', encoding='utf-8') as f:
    f.write(new_content)

print(f"\ncustomer_master.js updated:")
print(f"  Single-site records with postal_code added : {pins_added}")
print(f"  Multi-site locations processed (net added) : {locs_added}")
print("Done.")
