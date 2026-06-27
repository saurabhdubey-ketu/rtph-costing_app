#!/usr/bin/env python3
"""Build master import templates (one sheet per master) — pre-seeded with real
Indus Belts grades + values verified against the source sheets. Headers match
the Prisma model field names so the team's importer maps 1:1."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

HDR = Font(bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=13, color="305496")
NOTE = Font(italic=True, color="606060", size=9)
thin = Side(style="thin", color="D0D0D0")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def sheet(name, headers, rows, notes=""):
    ws = wb.create_sheet(name)
    if notes:
        ws["A1"] = notes; ws["A1"].font = NOTE
        start = 2
    else:
        start = 1
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=start, column=j, value=h); c.font = HDR; c.fill = HDRFILL; c.border = BORD
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(rows, start+1):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v); c.border = BORD
    ws.freeze_panes = ws.cell(row=start+1, column=1)
    for j, h in enumerate(headers, 1):
        w = max(12, min(30, max([len(str(h))] + [len(str(r[j-1])) for r in rows if j-1 < len(r)]) + 2))
        ws.column_dimensions[ws.cell(row=start, column=j).column_letter].width = w
    return ws

# ---------- Instructions ----------
ins = wb.active; ins.title = "READ ME FIRST"
ins["A1"] = "BELT COSTING — MASTER IMPORT TEMPLATES"; ins["A1"].font = TITLE
lines = [
 "",
 "HOW TO USE",
 "• One sheet per master. Row 1 = column headers (match the database fields). Fill rows below.",
 "• Rows already present are REAL seed data (Indus grades + values from your sheets). Edit/extend them.",
 "• YELLOW = prices/rates to keep current. Replace with your latest real numbers before go-live.",
 "• Codes must be UNIQUE within a sheet. Where a sheet references another (e.g. Fabric → supplierShortForm),",
 "  the referenced code must exist in the other sheet.",
 "• Multi-value cells use a pipe '|' separator (e.g. roles = TOP_COVER|BOTTOM_COVER).",
 "",
 "IMPORT ORDER (because of references)",
 "1 ProductType  2 BeltType  3 Supplier  4 FabricType  5 Fabric  6 BeltRating  7 NoOfPlyOption",
 "8 Compounds  9 CoverSkimCompatibility  10 Breaker  11 EdgeType  12 ReelPackingType",
 "13 WidthOption  14 FreightZone  15 CostOfProduction  16 Currency  17 StandardReference  18 Customer",
 "",
 "KEY RULES (from the spec)",
 "• A compound can hold MANY roles (cross-use). Skim is ONE list; tag skimUsage = FABRIC / STEEL_BREAKER / BOTH.",
 "• priceSource MANUAL now; becomes RAVASCO_COMPUTED after integration (price then comes from the recipe).",
 "• CoverSkimCompatibility drives the 'recommended skim' list. matchLevel FAMILY (common) or COMPOUND (unique, e.g. UHR=EPDM).",
 "• EdgeType.widthWastageMm: Cut Edge = 30, Moulded = 0 (drives width wastage).",
]
for i, t in enumerate(lines, 3):
    ins.cell(row=i, column=1, value=t).font = Font(bold=t.isupper() and t.strip()!="", size=10) if t.strip() else Font()
ins.column_dimensions["A"].width = 110

# ---------- Compounds ----------
sheet("Compounds",
 ["code","name","roles","chemicalCategory","gradeFamily","polymerBase","specificGravity","pricePerKg",
  "priceSource","skimUsage","tensileMpa","elongationPct","abrasionMm3","maxTempC","brandLine","standardRefIds","status","notes"],
 [
  ["M-24","M-24 General Purpose Cover","TOP_COVER|BOTTOM_COVER","M24","GP","NR/SBR",1.18,80,"MANUAL","","24","450","150","","Super Brute","IS 1891 M24","ACTIVE",""],
  ["M-15","M-15 General Purpose Cover","TOP_COVER|BOTTOM_COVER","M15","GP","NR/SBR",1.18,78,"MANUAL","","15","350","200","","Super Brute","IS 1891 M15","ACTIVE",""],
  ["N-17","N-17 Cover","TOP_COVER|BOTTOM_COVER","N17","GP","NR/SBR",1.17,76,"MANUAL","","17","400","200","","Super Brute","IS-N-17","ACTIVE",""],
  ["SAR","Super Abrasion Resistant Cover","TOP_COVER|BOTTOM_COVER","SAR","AR","NR/SBR/BR",1.16,110,"MANUAL","","17","400","70","","Super Brute","","ACTIVE",""],
  ["DIN-X","DIN X Abrasion Cover","TOP_COVER|BOTTOM_COVER","DINX","AR","NR/SBR",1.16,105,"MANUAL","","25","450","120","","Super Brute","DIN X","ACTIVE",""],
  ["HR-T1","Heat Resistant 125C Cover","TOP_COVER|BOTTOM_COVER","HR","HR","SBR",1.25,140,"MANUAL","","","","","125","Super Thermo","","ACTIVE",""],
  ["SHR-T2","Super Heat Resistant 150C Cover","TOP_COVER|BOTTOM_COVER","SHR","HR","SBR",1.25,150,"MANUAL","","","","","150","Super Thermo","","ACTIVE",""],
  ["UHR","Ultra Heat Resistant 200C Cover","TOP_COVER","UHR","HR","EPDM",1.20,165,"MANUAL","","","","","200","Super Thermo","","ACTIVE",""],
  ["FR","Fire Resistant Cover","TOP_COVER|BOTTOM_COVER","FR","FR","SBR/CR",1.35,130,"MANUAL","","","","","","Super Blaze","ISO 340","ACTIVE",""],
  ["OR-M","Oil Resistant (Moderate) Cover","TOP_COVER|BOTTOM_COVER","OR","OR","NBR",1.25,145,"MANUAL","","","","","","Super Slick","","ACTIVE",""],
  ["SK-FAB","General Purpose Fabric Skim","SKIM","","GP","NR/SBR",1.18,65,"MANUAL","FABRIC","","","","","","","ACTIVE","Carcass skim (the selectable skim)"],
  ["NN-III","Breaker Skim NN-III","SKIM","","GP","NR/SBR",1.22,65,"MANUAL","STEEL_BREAKER","","","","","","","ACTIVE",""],
  ["STBR-III","Breaker Skim STBR-III","SKIM","","GP","",1.25,88,"MANUAL","STEEL_BREAKER","","","","","","","ACTIVE",""],
  ["HR-SKIM","Heat Resistant Skim","SKIM","","HR","SBR",1.25,95,"MANUAL","BOTH","","","","","","","ACTIVE",""],
  ["UHR-SKIM","Ultra HR Skim (EPDM)","SKIM","","HR","EPDM",1.22,120,"MANUAL","BOTH","","","","","","","ACTIVE","EPDM — only bonds to UHR/SUHR covers"],
  ["SWF","Sidewall Compound","SIDEWALL","","GP","",1.25,150,"MANUAL","","","","","","","","ACTIVE",""],
  ["CLEAT-GP","Cleat Compound GP","CLEAT","","GP","",1.15,150,"MANUAL","","","","","","","","ACTIVE",""],
  ["BLINKER-GP","Blinker Compound","BLINKER","","GP","",1.18,150,"MANUAL","","","","","","","","ACTIVE",""],
  ["RB-SOL","RB Bonding Solution","SOLUTION","","GP","",0.85,1000,"MANUAL","","","","","","","","ACTIVE","Made or bought"],
  ["HARDENER","Standard Hardener","HARDENER","","OTHER","",1.00,750,"MANUAL","","","","","","","","ACTIVE","Bought-out"],
 ],
 "COMPOUND MASTER — the core. One row per compound. roles can be multiple (pipe-separated). skimUsage only for SKIM role.")

# ---------- CoverSkimCompatibility ----------
sheet("CoverSkimCompatibility",
 ["skimCode","matchLevel","coverGradeFamily","coverCompoundCode","isDefault","notes"],
 [
  ["SK-FAB","FAMILY","GP","","TRUE","Standard GP skim"],
  ["SK-FAB","FAMILY","AR","","TRUE","Standard AR skim"],
  ["HR-SKIM","FAMILY","HR","","TRUE","SBR heat skim for HR/SHR covers"],
  ["UHR-SKIM","COMPOUND","","UHR","TRUE","EPDM skim — only for UHR (won't bond to SBR covers)"],
 ],
 "COVER↔SKIM MATCH — which skim is recommended for which cover. FAMILY = whole grade family; COMPOUND = one specific cover.")

# ---------- FabricType ----------
sheet("FabricType",
 ["code","name","defaultPricePerKg","lengthWastagePct"],
 [["NN","Nylon-Nylon",300,0.04],["EP","Polyester-Polyester",265,0.03],["EE","Polyester-Polyester (EE)",210,0.03],
  ["EN","Polyester-Nylon",242,0.03],["PP","Polyamide (PP)",400,0.05]],
 "FABRIC TYPE — family + fallback price + length wastage %.")

# ---------- Fabric ----------
sheet("Fabric",
 ["code","fabricType","perPlyRating","supplierShortForm","gsm","thicknessMm","pricePerKg","supplierMaterialCode","internalCode"],
 [["EP-200-MIT","EP",200,"MIT",660,0.95,265,"EP-200",""],
  ["EP-158-MIT","EP",158,"MIT",570,0.80,265,"EP-158",""],
  ["NN-200-MIT","NN",200,"MIT",590,0.95,300,"NN-200",""],
  ["EE-167-SRF","EE",167,"SRF",720,0.95,210,"EE-167",""]],
 "FABRIC (per-ply, the buyable item). pricePerKg blank = use FabricType default. Multi-supplier supported.")

# ---------- BeltRating ----------
sheet("BeltRating",
 ["code","fabricType","totalBreakingStrength","noOfPly","perPlyRating","nominalCarcassThicknessMm","interPlyThicknessMm"],
 [["EP-1000/5","EP",1000,5,200,6.8,0.55],["EP-630/4","EP",630,4,158,5.2,0.50],
  ["NN-1000/5","NN",1000,5,200,7.0,0.55],["EE-500/3","EE",500,3,167,4.0,0.50]],
 "BELT RATING — the shorthand (EP-1000/5). Decodes to plies × per-ply fabric + carcass thickness.")

# ---------- Supplier ----------
sheet("Supplier",
 ["name","shortForm","supplierCode","location","locationCode"],
 [["Madura Industrial Textiles Limited","MIT","1","Dadra Unit","2"],
  ["SRF Limited","SRF","2","Gummidipoondi","1"]],
 "FABRIC/BREAKER SUPPLIER.")

# ---------- Breaker ----------
sheet("Breaker",
 ["code","breakerType","supplierShortForm","gsm","thicknessMm","noOfPlyDefault","pricePerKg","defaultMount","supplierMaterialCode"],
 [["BRK-TOP","Regular Fabric Breaker","MIT",140,0.30,1,400,"NON_FLOATING","MCO16"],
  ["BRK-BOT","Steel Breaker","MIT",650,1.50,1,550,"NON_FLOATING","MCO25"],
  ["BRK-CORD","Cord Breaker","MIT",1270,1.45,1,520,"FLOATING",""]],
 "BREAKER MASTER. defaultMount FLOATING / NON_FLOATING (team's good catch).")

# ---------- EdgeType ----------
sheet("EdgeType",
 ["code","name","widthWastageMm"],
 [["CUT_EDGE","Cut Edge",30],["MOULDED","Moulded",0],["VULCANISED","Vulcanised",0]],
 "EDGE TYPE — widthWastageMm drives width wastage (Cut Edge 30, Moulded 0). Fixes the wastage bug.")

# ---------- ReelPackingType ----------
sheet("ReelPackingType",
 ["code","name","packingCostPerMeter","appliesTo"],
 [["SINGLE_ROLL","Single Roll",0,"REEL"],["SWING_ROLL","Swing Roll",0,"REEL"],["CASSETTE","Cassette Roll",0,"REEL"],
  ["HDPE","HDPE Packing",4,"PACKING"],["WOODEN_CRATE","Wooden Crate",12,"PACKING"],["STEEL_CRATE","Steel Crate",200,"PACKING"]],
 "REEL / PACKING TYPE.")

# ---------- WidthOption ----------
sheet("WidthOption",
 ["widthMm","widthInch","label"],
 [[500,19.685,"500 mm"],[650,25.591,"650 mm"],[800,31.496,"800 mm"],[1000,39.370,"1000 mm"],
  [1200,47.244,"1200 mm"],[1400,55.118,"1400 mm"],[1600,62.992,"1600 mm"],[1800,70.866,"1800 mm"],[2000,78.740,"2000 mm"]],
 "WIDTH MASTER — width is chosen from this list (avoids manual entry errors). Add custom widths as rows.")

# ---------- NoOfPlyOption ----------
sheet("NoOfPlyOption", ["value"], [[i] for i in range(1,9)], "NO-OF-PLY MASTER (1–8).")

# ---------- FreightZone ----------
sheet("FreightZone",
 ["state","stateCode","city","freightRate","costType"],
 [["Jharkhand","JH","Jamshedpur",5.5,"KG"],["Gujarat","GJ","Surat",3.0,"KG"],
  ["Tamil Nadu","TN","Chennai",7.32,"KG"],["West Bengal","WB","Kolkata",9.89,"KG"],
  ["Maharashtra","MH","Mumbai",4.82,"KG"]],
 "FREIGHT ZONE — rate per destination. costType KG / SQMTR / RM.")

# ---------- CostOfProduction ----------
sheet("CostOfProduction",
 ["beltTypeCode","ratePerKg","effectiveFrom"],
 [["MULTIPLY",20,"2026-04-01"],["MULTIPLY_BREAKER",22,"2026-04-01"],["CHEVRON",30,"2026-04-01"],
  ["STEEP_ANGLE",30,"2026-04-01"],["SIDEWALL",30,"2026-04-01"]],
 "COST OF PRODUCTION — rate per belt type (locked: varies by belt type).")

# ---------- Currency ----------
sheet("Currency", ["code","name","exchangeRateToInr"],
 [["INR","Indian Rupee",1],["USD","US Dollar",83.50],["EUR","Euro",90.00]],
 "CURRENCY + exchange rate to INR.")

# ---------- StandardReference ----------
sheet("StandardReference", ["code","title","organization","region"],
 [["IS 1891 M24","IS 1891 Part 1 Grade M24","IS","India"],["IS-N-17","IS 1891 Grade N17","IS","India"],
  ["ISO 340","ISO 340 Flame Retardant","ISO","International"],["DIN X","DIN 22102 Grade X","DIN","EU"],
  ["MSHA 2G","MSHA Part 18 2G","MSHA","USA"]],
 "STANDARD REFERENCE (datasheets).")

# ---------- ProductType ----------
sheet("ProductType", ["code","name"],
 [["CB","Conveyor Belt"],["VB","V-Belt"],["TB","Transmission Belt"],["RS","Rubber Sheet"]],
 "PRODUCT TYPE (top level).")

# ---------- BeltType ----------
sheet("BeltType",
 ["code","name","productTypeCode","calcFamily","lengthRule","components","ravascoSubCategoryCode","isV1"],
 [
  ["MULTIPLY","Multi-Ply Textile","CB","BASE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["MULTIPLY_BREAKER","Multi-Ply with Breaker","CB","BASE_BREAKER","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|BREAKER_TOP|BREAKER_TOP_SKIM|BREAKER_BOTTOM|BREAKER_BOTTOM_SKIM","TEXTILE","TRUE"],
  ["BAG_DIVERTER","Bag Diverter","CB","BASE_BREAKER","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|BREAKER_TOP|BREAKER_TOP_SKIM","TEXTILE","TRUE"],
  ["BUCKET_ELEVATOR","Bucket Elevator","CB","BASE_BREAKER","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|BREAKER_TOP|BREAKER_TOP_SKIM","TEXTILE","TRUE"],
  ["STRAIGHT_WARP","Straight Warp","CB","BASE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["ROUGH_TOP","Rough Top","CB","BASE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["WAVY_TOP","Wavy Top","CB","BASE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["ENDLESS","Endless","CB","BASE","ENDLESS","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["JOINTLESS","Jointless","CB","BASE","ENDLESS","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","TRUE"],
  ["CHEVRON","Chevron","CB","BASE_CLEAT","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|CLEAT","TEXTILE","TRUE"],
  ["STEEP_ANGLE","Steep Angle","CB","FULL","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|CLEAT|SIDEWALL|BLINKER|SOLUTION|HARDENER","TEXTILE","TRUE"],
  ["SIDEWALL","Sidewall","CB","SIDEWALL","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM|SIDEWALL|SOLUTION|HARDENER","TEXTILE","TRUE"],
  ["PIPE","Pipe","CB","BASE_PIPE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","FALSE"],
  ["PAPER_REEL","Paper Reel","CB","BASE","OPEN_END","TOP_COVER|BOTTOM_COVER|FABRIC|SKIM","TEXTILE","FALSE"],
 ],
 "BELT TYPE — declares calcFamily + components shown per type. isV1 FALSE = Pipe/Paper Reel (v2).")

# ---------- Customer ----------
sheet("Customer",
 ["name","gstin","defaultDestination","defaultGpPct","defaultCurrency","discountTier"],
 [["(example) Tata Steel Ltd","27AAACT1234A1Z5","Jamshedpur",0.30,"INR","A"]],
 "CUSTOMER — header defaults for quotations.")

out = os.path.join(os.path.dirname(__file__), "Belt-Costing-Master-Import-Templates.xlsx")
wb.save(out)
print("Saved:", out, "with", len(wb.sheetnames), "sheets")
print("Sheets:", ", ".join(wb.sheetnames))
