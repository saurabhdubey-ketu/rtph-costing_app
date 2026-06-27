#!/usr/bin/env python3
"""Build the corrected, master-driven Multiply-with-Breaker costing workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ---- styles ----
HDR = Font(bold=True, color="FFFFFF", size=11)
HDRFILL = PatternFill("solid", fgColor="305496")
SUB = Font(bold=True, color="000000", size=10)
SUBFILL = PatternFill("solid", fgColor="D9E1F2")
BLUE = Font(color="0000FF")          # inputs
GREEN = Font(color="008000")         # cross-sheet links
BLACK = Font(color="000000")         # in-sheet formulas
YEL = PatternFill("solid", fgColor="FFFF00")
NOTE = Font(italic=True, color="808080", size=9)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '#,##0.00;(#,##0.00);"-"'
KG = '#,##0.00'

wb = Workbook()

# ===================== MASTERS SHEET =====================
m = wb.active
m.title = "Masters"
m.sheet_properties.tabColor = "70AD47"

def hdr(ws, cell, text):
    ws[cell] = text; ws[cell].font = HDR; ws[cell].fill = HDRFILL

def sub(ws, cell, text):
    ws[cell] = text; ws[cell].font = SUB; ws[cell].fill = SUBFILL

m["A1"] = "MASTERS — single source of all rates. Change a rate here; every costing updates."
m["A1"].font = Font(bold=True, size=12, color="305496")

# --- Compounds ---
sub(m, "A3", "COMPOUND MASTER")
heads = ["Code", "Name", "Role", "Grade Family", "SG", "Price/kg"]
for i, h in enumerate(heads):
    c = m.cell(row=4, column=1+i, value=h); c.font = SUB; c.fill = SUBFILL; c.border = BORD
compounds = [
    ["M-24", "M-24 General Purpose Cover", "Cover", "GP", 1.18, 80],
    ["M-15", "M-15 General Purpose Cover", "Cover", "GP", 1.18, 78],
    ["SAR",  "Super Abrasion Resistant Cover", "Cover", "AR", 1.16, 110],
    ["HR-T1","Heat Resistant 125C Cover", "Cover", "HR", 1.25, 140],
    ["UHR",  "Ultra Heat Resistant 200C (EPDM)", "Cover", "HR", 1.20, 165],
    ["SK-FAB","General Purpose Fabric Skim", "Skim", "GP", 1.18, 65],
    ["NN-III","Breaker Skim NN-III", "Skim", "GP", 1.22, 65],
    ["STBR-III","Breaker Skim STBR-III", "Skim", "GP", 1.25, 88],
    ["HR-SKIM","Heat Resistant Skim", "Skim", "HR", 1.25, 95],
    ["UHR-SKIM","Ultra Heat Resistant Skim (EPDM)", "Skim", "HR", 1.22, 120],
]
r = 5
for row in compounds:
    for i, v in enumerate(row):
        c = m.cell(row=r, column=1+i, value=v); c.border = BORD
        if i == 5: c.number_format = MONEY
    r += 1
COMP_RANGE = f"Masters!$A$5:$F${r-1}"
COMP_CODES = f"Masters!$A$5:$A${r-1}"
# yellow-flag the price column as "update me"
for rr in range(5, r):
    m.cell(row=rr, column=6).fill = YEL

# --- Fabric / Belt Rating ---
fr0 = r + 1
sub(m, f"A{fr0}", "FABRIC / BELT RATING MASTER")
heads = ["Belt Rating", "Fabric Type", "Total Strength", "No of Ply", "Per-Ply Rating", "GSM", "Carcass Thk (mm)", "Price/kg"]
for i, h in enumerate(heads):
    c = m.cell(row=fr0+1, column=1+i, value=h); c.font = SUB; c.fill = SUBFILL; c.border = BORD
fabrics = [
    ["EP-1000/5", "EP", 1000, 5, 200, 660, 6.8, 240],
    ["EP-630/4",  "EP", 630, 4, 158, 570, 5.2, 240],
    ["NN-1000/5", "NN", 1000, 5, 200, 590, 7.0, 300],
    ["EE-500/3",  "EE", 500, 3, 167, 720, 4.0, 210],
]
rr = fr0 + 2
for row in fabrics:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
        if i == 7: c.number_format = MONEY
    rr += 1
FAB_RANGE = f"Masters!$A${fr0+2}:$H${rr-1}"
FAB_CODES = f"Masters!$A${fr0+2}:$A${rr-1}"

# --- Breaker ---
br0 = rr + 1
sub(m, f"A{br0}", "BREAKER MASTER")
heads = ["Code", "Breaker Type", "GSM", "Thickness (mm)", "No of Ply", "Price/kg"]
for i, h in enumerate(heads):
    c = m.cell(row=br0+1, column=1+i, value=h); c.font = SUB; c.fill = SUBFILL; c.border = BORD
breakers = [
    ["BRK-TOP", "Regular Fabric Breaker", 140, 0.3, 1, 400],
    ["BRK-BOT", "Steel Breaker", 650, 1.5, 1, 550],
    ["BRK-CORD", "Cord Breaker", 1270, 1.45, 1, 520],
]
rr = br0 + 2
for row in breakers:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
        if i == 5: c.number_format = MONEY
    rr += 1
BRK_RANGE = f"Masters!$A${br0+2}:$F${rr-1}"
BRK_CODES = f"Masters!$A${br0+2}:$A${rr-1}"

# --- Edge ---
ed0 = rr + 1
sub(m, f"A{ed0}", "EDGE MASTER")
m.cell(row=ed0+1, column=1, value="Edge Type").font = SUB
m.cell(row=ed0+1, column=2, value="Width Wastage (mm)").font = SUB
edges = [["Cut Edge", 30], ["Moulded", 0], ["Vulcanised", 0]]
rr = ed0 + 2
for row in edges:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
    rr += 1
EDGE_RANGE = f"Masters!$A${ed0+2}:$B${rr-1}"
EDGE_CODES = f"Masters!$A${ed0+2}:$A${rr-1}"

# --- Freight ---
frt0 = rr + 1
sub(m, f"A{frt0}", "FREIGHT MASTER")
m.cell(row=frt0+1, column=1, value="Destination").font = SUB
m.cell(row=frt0+1, column=2, value="Rate").font = SUB
m.cell(row=frt0+1, column=3, value="Cost Type").font = SUB
freights = [["Jamshedpur", 5.5, "KG"], ["Surat", 3.0, "KG"], ["Chennai", 7.32, "KG"], ["Kolkata", 9.89, "KG"]]
rr = frt0 + 2
for row in freights:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
    rr += 1
FRT_RANGE = f"Masters!$A${frt0+2}:$C${rr-1}"
FRT_CODES = f"Masters!$A${frt0+2}:$A${rr-1}"

# --- Packing ---
pk0 = rr + 1
sub(m, f"A{pk0}", "PACKING MASTER")
m.cell(row=pk0+1, column=1, value="Packing Type").font = SUB
m.cell(row=pk0+1, column=2, value="Cost/m").font = SUB
packs = [["HDPE Packing", 4], ["Wooden Crate", 12], ["Steel Crate", 200]]
rr = pk0 + 2
for row in packs:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
    rr += 1
PK_RANGE = f"Masters!$A${pk0+2}:$B${rr-1}"
PK_CODES = f"Masters!$A${pk0+2}:$A${rr-1}"

# --- Cost of production ---
cp0 = rr + 1
sub(m, f"A{cp0}", "COST-OF-PRODUCTION MASTER (per belt type)")
m.cell(row=cp0+1, column=1, value="Belt Type").font = SUB
m.cell(row=cp0+1, column=2, value="Rate/kg").font = SUB
cprod = [["Multiply with Breaker", 22], ["Steep Angle", 30], ["Chevron", 30], ["Sidewall", 30]]
rr = cp0 + 2
for row in cprod:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
    rr += 1
CP_RANGE = f"Masters!$A${cp0+2}:$B${rr-1}"
CP_CODES = f"Masters!$A${cp0+2}:$A${rr-1}"

# --- Assumptions ---
as0 = rr + 1
sub(m, f"A{as0}", "ASSUMPTIONS")
asmp = [["Open-End Length Wastage %", 0.03], ["Endless Splice Allowance (m)", 3]]
rr = as0 + 1
for row in asmp:
    m.cell(row=rr, column=1, value=row[0]).border = BORD
    c = m.cell(row=rr, column=2, value=row[1]); c.border = BORD; c.fill = YEL
    rr += 1
OPENWASTE = f"Masters!$B${as0+1}"
SPLICE = f"Masters!$B${as0+2}"

# --- skim recommendation helper (cover family -> recommended skim) ---
sk0 = rr + 1
sub(m, f"A{sk0}", "COVER↔SKIM RECOMMENDED MATCH")
m.cell(row=sk0+1, column=1, value="Cover Family").font = SUB
m.cell(row=sk0+1, column=2, value="Recommended Skim").font = SUB
skmatch = [["GP", "SK-FAB"], ["AR", "SK-FAB"], ["HR", "HR-SKIM"]]
rr = sk0 + 2
for row in skmatch:
    for i, v in enumerate(row):
        c = m.cell(row=rr, column=1+i, value=v); c.border = BORD
    rr += 1
SKMATCH_RANGE = f"Masters!$A${sk0+2}:$B${rr-1}"

for col, w in {"A": 30, "B": 26, "C": 16, "D": 16, "E": 14, "F": 12, "G": 16, "H": 12}.items():
    m.column_dimensions[col].width = w

# ===================== COSTING SHEET =====================
c = wb.create_sheet("Multiply with Breaker")
c.sheet_properties.tabColor = "305496"
c["A1"] = "MULTIPLY CONVEYOR BELT WITH BREAKER — COSTING"
c["A1"].font = Font(bold=True, size=13, color="305496")
c["A2"] = "Corrected & master-driven. Blue = you type. Green = pulled from Masters. Black = calculated."
c["A2"].font = NOTE

def lbl(row, text):
    c.cell(row=row, column=1, value=text).font = Font(size=10)
def inp(row, val, dv=None):
    cell = c.cell(row=row, column=2, value=val); cell.font = BLUE; cell.fill = PatternFill("solid", fgColor="FFF2CC"); cell.border = BORD
    return cell
def lnk(row, formula, fmt=None):
    cell = c.cell(row=row, column=2, value=formula); cell.font = GREEN; cell.border = BORD
    if fmt: cell.number_format = fmt
    return cell

hdr(c, "A4", "INPUTS — BELT CONFIGURATION"); c["B4"].fill = HDRFILL
lbl(5, "Width (mm)"); inp(5, 1000)
lbl(6, "Belt Rating"); inp(6, "EP-1000/5")
lbl(7, "Fabric Type"); lnk(7, "=VLOOKUP($B$6,%s,2,0)" % FAB_RANGE)
lbl(8, "Total Strength (N/mm)"); lnk(8, "=VLOOKUP($B$6,%s,3,0)" % FAB_RANGE)
lbl(9, "No of Ply"); lnk(9, "=VLOOKUP($B$6,%s,4,0)" % FAB_RANGE)
lbl(10, "Carcass Thickness (mm)"); lnk(10, "=VLOOKUP($B$6,%s,7,0)" % FAB_RANGE)
lbl(11, "Fabric GSM"); lnk(11, "=VLOOKUP($B$6,%s,6,0)" % FAB_RANGE)
lbl(12, "Fabric Price/kg"); lnk(12, "=VLOOKUP($B$6,%s,8,0)" % FAB_RANGE, MONEY)
lbl(13, "Top Cover (mm)"); inp(13, 5)
lbl(14, "Bottom Cover (mm)"); inp(14, 2)
lbl(15, "Top Cover Compound"); inp(15, "M-24")
lbl(16, "Bottom Cover Compound"); inp(16, "M-24")
cell = c.cell(row=17, column=1, value="Fabric Skim Compound  ◄ NEW"); cell.font = Font(size=10, bold=True, color="C00000")
inp(17, "SK-FAB")
lbl(18, "Edge Type"); inp(18, "Cut Edge")
lbl(19, "Open End / Endless"); inp(19, "Open End")

hdr(c, "A21", "INPUTS — BREAKERS"); c["B21"].fill = HDRFILL
lbl(22, "Breaker on Top? (Yes/No)"); inp(22, "Yes")
lbl(23, "Breaker Top Code"); inp(23, "BRK-TOP")
lbl(24, "Breaker Top Skim Compound"); inp(24, "NN-III")
lbl(25, "Breaker Top Skim Thickness (mm)"); inp(25, 0.6)
lbl(26, "Breaker on Bottom? (Yes/No)"); inp(26, "Yes")
lbl(27, "Breaker Bottom Code"); inp(27, "BRK-BOT")
lbl(28, "Breaker Bottom Skim Compound"); inp(28, "STBR-III")
lbl(29, "Breaker Bottom Skim Thickness (mm)"); inp(29, 3)

hdr(c, "A31", "INPUTS — COMMERCIAL"); c["B31"].fill = HDRFILL
lbl(32, "Quantity per Roll (m)"); inp(32, 200)
lbl(33, "No of Rolls"); inp(33, 5)
lbl(34, "Total Quantity (m)"); cell = c.cell(row=34, column=2, value="=$B$32*$B$33"); cell.font = BLACK; cell.border = BORD
lbl(35, "Cost of Production rate /kg"); lnk(35, '=VLOOKUP("Multiply with Breaker",%s,2,0)' % CP_RANGE, MONEY)
lbl(36, "Packing Type"); inp(36, "HDPE Packing")
lbl(37, "Packing Cost/m"); lnk(37, "=VLOOKUP($B$36,%s,2,0)" % PK_RANGE, MONEY)
lbl(38, "Freight Destination"); inp(38, "Jamshedpur")
lbl(39, "Freight Rate"); lnk(39, "=VLOOKUP($B$38,%s,2,0)" % FRT_RANGE, MONEY)
lbl(40, "Freight Cost Type"); lnk(40, "=VLOOKUP($B$38,%s,3,0)" % FRT_RANGE)
lbl(41, "GP %"); inp(41, 0.35); c["B41"].number_format = "0%"
lbl(42, "Discount %"); inp(42, 0); c["B42"].number_format = "0%"

# skim match check
hdr(c, "A44", "SKIM ↔ COVER MATCH CHECK"); c["B44"].fill = HDRFILL
lbl(45, "Top Cover Grade Family"); lnk(45, "=VLOOKUP($B$15,%s,4,0)" % COMP_RANGE)
lbl(46, "Selected Skim Grade Family"); lnk(46, "=VLOOKUP($B$17,%s,4,0)" % COMP_RANGE)
lbl(47, "Recommended Skim for this Cover"); lnk(47, "=IFERROR(VLOOKUP($B$45,%s,2,0),\"(define in Masters)\")" % SKMATCH_RANGE)
lbl(48, "Match Status")
cell = c.cell(row=48, column=2, value='=IF($B$45=$B$46,"OK — skim matches cover family","⚠ WARNING: skim family does not match cover — confirm bonding")')
cell.font = Font(bold=True); cell.border = BORD

# derived helpers
hdr(c, "A50", "DERIVED"); c["B50"].fill = HDRFILL
c.cell(row=51, column=1, value="Effective Width W_eff (m)  ◄ FIX C8 (edge-driven)").font = Font(size=10)
c.cell(row=51, column=2, value="=($B$5+VLOOKUP($B$18,%s,2,0))/1000" % EDGE_RANGE).font = BLACK; c["B51"].border = BORD
c.cell(row=52, column=1, value="Effective Length L_eff (m)  ◄ FIX X2 (rule-driven)").font = Font(size=10)
c.cell(row=52, column=2, value="=IF($B$19=\"Open End\",$B$34*(1+%s),$B$34+%s)" % (OPENWASTE, SPLICE)).font = BLACK; c["B52"].border = BORD
c.cell(row=53, column=1, value="Belt wt without breaker (helper)").font = NOTE
c.cell(row=53, column=2, value="=$B$51*($B$10+$B$13+$B$14)*VLOOKUP($B$15,%s,5,0)*$B$52" % COMP_RANGE).font = NOTE; c["B53"].number_format = KG

# component table
hdr(c, "A56", "WEIGHT & COST"); c["B56"].fill = HDRFILL; c["C56"].fill = HDRFILL; c["D56"].fill = HDRFILL
for i, h in enumerate(["Component", "Weight (kg)", "Price/kg", "Cost (₹)"]):
    cell = c.cell(row=57, column=1+i, value=h); cell.font = SUB; cell.fill = SUBFILL; cell.border = BORD

def comp(row, name, wformula, pformula):
    c.cell(row=row, column=1, value=name).border = BORD
    wc = c.cell(row=row, column=2, value=wformula); wc.font = BLACK; wc.number_format = KG; wc.border = BORD
    pc = c.cell(row=row, column=3, value=pformula); pc.font = GREEN; pc.number_format = MONEY; pc.border = BORD
    dc = c.cell(row=row, column=4, value=f"=B{row}*C{row}"); dc.font = BLACK; dc.number_format = MONEY; dc.border = BORD

comp(58, "Top Cover", "=$B$51*$B$13*VLOOKUP($B$15,%s,5,0)*$B$52" % COMP_RANGE, "=VLOOKUP($B$15,%s,6,0)" % COMP_RANGE)
comp(59, "Bottom Cover", "=$B$51*$B$14*VLOOKUP($B$16,%s,5,0)*$B$52" % COMP_RANGE, "=VLOOKUP($B$16,%s,6,0)" % COMP_RANGE)
comp(60, "Carcass Fabric", "=$B$51*($B$11/1000)*$B$52*$B$9", "=$B$12")
# fabric skim = helper(B53) - top - bottom - fabric
c.cell(row=61, column=1, value="Fabric Skim  ◄ now selectable").border = BORD
wc = c.cell(row=61, column=2, value="=$B$53-B58-B59-B60"); wc.font = BLACK; wc.number_format = KG; wc.border = BORD
pc = c.cell(row=61, column=3, value="=VLOOKUP($B$17,%s,6,0)" % COMP_RANGE); pc.font = GREEN; pc.number_format = MONEY; pc.border = BORD
dc = c.cell(row=61, column=4, value="=B61*C61"); dc.font = BLACK; dc.number_format = MONEY; dc.border = BORD
comp(62, "Breaker Top", "=IF($B$22=\"Yes\",$B$51*(VLOOKUP($B$23,%s,3,0)/1000)*$B$52*VLOOKUP($B$23,%s,5,0),0)" % (BRK_RANGE, BRK_RANGE), "=IF($B$22=\"Yes\",VLOOKUP($B$23,%s,6,0),0)" % BRK_RANGE)
comp(63, "Breaker Top Skim", "=IF($B$22=\"Yes\",$B$51*$B$25*VLOOKUP($B$24,%s,5,0)*$B$52,0)" % COMP_RANGE, "=IF($B$22=\"Yes\",VLOOKUP($B$24,%s,6,0),0)" % COMP_RANGE)
comp(64, "Breaker Bottom", "=IF($B$26=\"Yes\",$B$51*(VLOOKUP($B$27,%s,3,0)/1000)*$B$52*VLOOKUP($B$27,%s,5,0),0)" % (BRK_RANGE, BRK_RANGE), "=IF($B$26=\"Yes\",VLOOKUP($B$27,%s,6,0),0)" % BRK_RANGE)
comp(65, "Breaker Bottom Skim", "=IF($B$26=\"Yes\",$B$51*$B$29*VLOOKUP($B$28,%s,5,0)*$B$52,0)" % COMP_RANGE, "=IF($B$26=\"Yes\",VLOOKUP($B$28,%s,6,0),0)" % COMP_RANGE)

c.cell(row=66, column=1, value="Total Belt Weight").font = SUB
tw = c.cell(row=66, column=2, value="=SUM(B58:B65)"); tw.font = SUB; tw.number_format = KG; tw.border = BORD
c.cell(row=66, column=3, value="Material Cost →").font = NOTE
mc = c.cell(row=66, column=4, value="=SUM(D58:D65)"); mc.font = SUB; mc.number_format = MONEY; mc.border = BORD

# other costs
c.cell(row=68, column=1, value="Cost of Production")
c.cell(row=68, column=4, value="=$B$35*$B$66").number_format = MONEY; c["D68"].border = BORD
c.cell(row=69, column=1, value="Packing Cost  ◄ FIX C2 (uses length)")
c.cell(row=69, column=4, value="=$B$37*$B$34").number_format = MONEY; c["D69"].border = BORD
c.cell(row=70, column=1, value="Freight Cost")
c.cell(row=70, column=4, value="=IF($B$40=\"KG\",$B$39*$B$66,IF($B$40=\"RM\",$B$39*$B$34,$B$39*($B$51*$B$52)))").number_format = MONEY; c["D70"].border = BORD
c.cell(row=71, column=1, value="TOTAL BELT COST").font = SUB
c.cell(row=71, column=4, value="=$D$66+$D$68+$D$69+$D$70").font = SUB; c["D71"].number_format = MONEY; c["D71"].border = BORD

# pricing
hdr(c, "A73", "PRICING"); c["B73"].fill = HDRFILL
c.cell(row=74, column=1, value="RMC per meter  ◄ FIX C1 (÷ length)")
c.cell(row=74, column=2, value="=$D$71/$B$34").number_format = MONEY; c["B74"].border = BORD
for i, h in enumerate(["", "Total (₹)", "Per Meter (₹)"]):
    cc = c.cell(row=76, column=1+i, value=h); cc.font = SUB; cc.fill = SUBFILL
c.cell(row=77, column=1, value="Standard (GP on full cost)")
c.cell(row=77, column=2, value="=$D$71+($D$71*$B$41)").number_format = MONEY; c["B77"].border = BORD
c.cell(row=77, column=3, value="=$B$77/$B$34").number_format = MONEY; c["C77"].border = BORD
c.cell(row=78, column=1, value="VD (GP on material only)")
c.cell(row=78, column=2, value="=$D$71+($D$66*$B$41)").number_format = MONEY; c["B78"].border = BORD
c.cell(row=78, column=3, value="=$B$78/$B$34").number_format = MONEY; c["C78"].border = BORD
c.cell(row=79, column=1, value="Standard Final (after discount)").font = SUB
c.cell(row=79, column=2, value="=$B$77*(1-$B$42)").number_format = MONEY; c["B79"].font = SUB; c["B79"].border = BORD
c.cell(row=79, column=3, value="=$B$79/$B$34").number_format = MONEY; c["C79"].font = SUB; c["C79"].border = BORD
c.cell(row=80, column=1, value="VD Final (after discount)").font = SUB
c.cell(row=80, column=2, value="=$B$78*(1-$B$42)").number_format = MONEY; c["B80"].font = SUB; c["B80"].border = BORD
c.cell(row=80, column=3, value="=$B$80/$B$34").number_format = MONEY; c["C80"].font = SUB; c["C80"].border = BORD
c.cell(row=82, column=1, value="No rounding applied (exact rate) — per Jitesh.").font = NOTE

c.column_dimensions["A"].width = 38
c.column_dimensions["B"].width = 18
c.column_dimensions["C"].width = 16
c.column_dimensions["D"].width = 16

# ---- data validations (dropdowns) ----
def add_dv(rng_cells, source):
    dv = DataValidation(type="list", formula1=source, allow_blank=True)
    c.add_data_validation(dv)
    for cell in rng_cells:
        dv.add(c[cell])
add_dv(["B6"], "=%s" % FAB_CODES)
add_dv(["B15", "B16"], "=%s" % COMP_CODES)
add_dv(["B17", "B24", "B28"], "=%s" % COMP_CODES)
add_dv(["B18"], "=%s" % EDGE_CODES)
add_dv(["B19"], '"Open End,Endless"')
add_dv(["B22", "B26"], '"Yes,No"')
add_dv(["B23", "B27"], "=%s" % BRK_CODES)
add_dv(["B36"], "=%s" % PK_CODES)
add_dv(["B38"], "=%s" % FRT_CODES)

# comments on key fixes
c["A17"].comment = Comment("NEW: Fabric skim compound is now selectable from the compound master (role=Skim). Was hardcoded at 65 in the old sheet.", "Spec")
c["A51"].comment = Comment("FIX C8: width wastage now comes from the Edge master (Cut Edge=30, Moulded=0), not hardcoded +30.", "Spec")
c["A69"].comment = Comment("FIX C2: packing cost uses total length (m), not pieces.", "Spec")
c["A74"].comment = Comment("FIX C1: RMC always divides total cost by length.", "Spec")

import os
out = os.path.join(os.path.dirname(__file__), "Corrected-Multiply-Belt-Costing.xlsx")
wb.save(out)
print("Saved:", out)
